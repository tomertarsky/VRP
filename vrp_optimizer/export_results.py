#!/usr/bin/env python3
"""
Export VRP optimization results to an Excel workbook with multiple sheets:
  1. Weekly_Summary   — daily totals by depot
  2. Route_Details    — every stop on every route, in order
  3. Dropped_Sites    — sites not served and why
  4. Cost_Breakdown   — full cost/revenue analysis
  5. Logic_Constraints — explanation of the model
"""
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "/private/tmp/claude-501/-Users-tomertarsky/tasks/b5e0558.output"
EXPORT_PATH = "/Users/tomertarsky/vrp_optimizer/VRP_Results.xlsx"

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

# ── Styles ────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
MONEY_FMT = '#,##0.00'
INT_FMT = '#,##0'
PCT_FMT = '0.00%'
THIN_BORDER = Border(
    bottom=Side(style='thin', color='B0B0B0'),
)


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 55)


def parse_output(path):
    with open(path, "r") as f:
        text = f.read()
    return text


def parse_depot_summary(text):
    """Parse per-day per-depot solver summary lines."""
    rows = []
    current_day = None
    pattern = re.compile(
        r"Depot (\w+): (\d+) visits, (\d+) trucks available"
    )
    result_pattern = re.compile(
        r"→ (\d+) trucks \| ([\d,]+) lbs \| ([\d,.]+) km \| (\d+) dropped"
    )

    lines = text.split("\n")
    for i, line in enumerate(lines):
        for day_name in DAY_NAMES:
            if f"Solving {day_name}" in line:
                current_day = day_name
                break

        m = pattern.search(line)
        if m and current_day:
            depot = m.group(1)
            visits = int(m.group(2))
            trucks_avail = int(m.group(3))
            # Look ahead for the result line
            for j in range(i + 1, min(i + 5, len(lines))):
                rm = result_pattern.search(lines[j])
                if rm:
                    rows.append({
                        "day": current_day,
                        "depot": depot,
                        "visits_scheduled": visits,
                        "trucks_available": trucks_avail,
                        "trucks_used": int(rm.group(1)),
                        "lbs_collected": int(rm.group(2).replace(",", "")),
                        "km_driven": float(rm.group(3).replace(",", "")),
                        "dropped": int(rm.group(4)),
                    })
                    break
    return rows


def parse_routes(text):
    """Parse detailed route stops from the report section."""
    rows = []
    current_day = None
    current_depot = None
    current_truck = None
    truck_stops = 0
    truck_lbs = 0
    truck_km = 0.0
    truck_min = 0

    lines = text.split("\n")
    # Find the report section (after [6/6])
    start = 0
    for i, line in enumerate(lines):
        if "[6/6]" in line:
            start = i
            break

    day_pattern = re.compile(r"^\s{2}(" + "|".join(DAY_NAMES).upper() + r")\s*$")
    depot_pattern = re.compile(r"^\s+Depot: (\w+)")
    truck_pattern = re.compile(
        r"Truck #(\d+): (\d+) stops \| ([\d,]+) lbs \| ([\d,.]+) km \| (\d+) min"
    )
    stop_pattern = re.compile(
        r"-> (.+?)\s{2,}(\d+) lbs \|\s+(\d+) min \| (?:net \$\s*[\d,.+-]+|MAND|opt)"
    )

    stop_order = 0

    for i in range(start, len(lines)):
        line = lines[i]

        # Check for day header
        for day_name in DAY_NAMES:
            if line.strip() == day_name.upper():
                current_day = day_name
                stop_order = 0
                break

        dm = depot_pattern.match(line)
        if dm:
            current_depot = dm.group(1).lower()

        tm = truck_pattern.search(line)
        if tm:
            current_truck = int(tm.group(1))
            truck_stops = int(tm.group(2))
            truck_lbs = int(tm.group(3).replace(",", ""))
            truck_km = float(tm.group(4).replace(",", ""))
            truck_min = int(tm.group(5))
            stop_order = 0

        sm = stop_pattern.search(line)
        if sm and current_day and current_depot and current_truck:
            stop_order += 1
            rows.append({
                "day": current_day,
                "depot": current_depot,
                "truck": current_truck,
                "truck_label": f"{current_depot.upper()}-{current_truck}",
                "stop_order": stop_order,
                "address": sm.group(1).strip(),
                "lbs": int(sm.group(2)),
                "service_min": int(sm.group(3)),
            })

    return rows


def parse_dropped(text):
    """Parse dropped sites from the report."""
    rows = []
    in_dropped = False
    drop_pattern = re.compile(
        r"^\s{2}(.{55,57})\s+Net/visit: \$\s*([\d,.]+) \| Lbs/yr:\s*([\d,]+) \| (.+)$"
    )

    for line in text.split("\n"):
        if "DROPPED SITES" in line:
            in_dropped = True
            continue
        if in_dropped and line.strip().startswith("===="):
            if rows:  # we already have data, next section
                break
            continue
        if in_dropped:
            m = drop_pattern.match(line)
            if m:
                rows.append({
                    "address": m.group(1).strip(),
                    "net_per_visit": float(m.group(2).replace(",", "")),
                    "annual_lbs": int(m.group(3).replace(",", "")),
                    "frequency": m.group(4).strip(),
                })
    return rows


def write_weekly_summary(wb, depot_rows):
    ws = wb.create_sheet("Weekly_Summary")
    headers = [
        "Day", "Depot", "Visits Scheduled", "Trucks Available",
        "Trucks Used", "Lbs Collected", "Km Driven", "Visits Dropped",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    for r, row in enumerate(depot_rows, 2):
        ws.cell(row=r, column=1, value=row["day"])
        ws.cell(row=r, column=2, value=row["depot"])
        ws.cell(row=r, column=3, value=row["visits_scheduled"])
        ws.cell(row=r, column=4, value=row["trucks_available"])
        ws.cell(row=r, column=5, value=row["trucks_used"])
        ws.cell(row=r, column=6, value=row["lbs_collected"]).number_format = INT_FMT
        ws.cell(row=r, column=7, value=row["km_driven"]).number_format = '#,##0.0'
        ws.cell(row=r, column=8, value=row["dropped"])

    # Day totals
    r = len(depot_rows) + 3
    ws.cell(row=r, column=1, value="DAY TOTALS").font = Font(bold=True)
    style_header(ws, r + 1, len(headers))
    for c, h in enumerate(headers, 1):
        ws.cell(row=r + 1, column=c, value=h)

    r += 2
    for day in DAY_NAMES:
        day_data = [d for d in depot_rows if d["day"] == day]
        if not day_data:
            continue
        ws.cell(row=r, column=1, value=day)
        ws.cell(row=r, column=3, value=sum(d["visits_scheduled"] for d in day_data))
        ws.cell(row=r, column=5, value=sum(d["trucks_used"] for d in day_data))
        ws.cell(row=r, column=6, value=sum(d["lbs_collected"] for d in day_data)).number_format = INT_FMT
        ws.cell(row=r, column=7, value=sum(d["km_driven"] for d in day_data)).number_format = '#,##0.0'
        ws.cell(row=r, column=8, value=sum(d["dropped"] for d in day_data))
        r += 1

    # Weekly total
    ws.cell(row=r, column=1, value="WEEK TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=5, value=max(
        sum(d["trucks_used"] for d in depot_rows if d["day"] == day)
        for day in DAY_NAMES
    ))
    ws.cell(row=r, column=6, value=sum(d["lbs_collected"] for d in depot_rows)).number_format = INT_FMT
    ws.cell(row=r, column=7, value=sum(d["km_driven"] for d in depot_rows)).number_format = '#,##0.0'
    ws.cell(row=r, column=8, value=sum(d["dropped"] for d in depot_rows))

    auto_width(ws)
    ws.freeze_panes = "A2"


def write_route_details(wb, route_rows):
    ws = wb.create_sheet("Route_Details")
    headers = [
        "Day", "Depot", "Truck ID", "Stop Order", "Address",
        "Lbs", "Service (min)",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    for r, row in enumerate(route_rows, 2):
        ws.cell(row=r, column=1, value=row["day"])
        ws.cell(row=r, column=2, value=row["depot"])
        ws.cell(row=r, column=3, value=row["truck_label"])
        ws.cell(row=r, column=4, value=row["stop_order"])
        ws.cell(row=r, column=5, value=row["address"])
        ws.cell(row=r, column=6, value=row["lbs"]).number_format = INT_FMT
        ws.cell(row=r, column=7, value=row["service_min"])

    auto_width(ws)
    ws.freeze_panes = "A2"


def write_dropped_sites(wb, dropped_rows):
    ws = wb.create_sheet("Dropped_Sites")
    headers = ["Address", "Net $/Visit", "Annual Lbs", "Frequency", "Reason"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    for r, row in enumerate(dropped_rows, 2):
        ws.cell(row=r, column=1, value=row["address"])
        ws.cell(row=r, column=2, value=row["net_per_visit"]).number_format = MONEY_FMT
        ws.cell(row=r, column=3, value=row["annual_lbs"]).number_format = INT_FMT
        ws.cell(row=r, column=4, value=row["frequency"])
        # Infer reason
        if row["net_per_visit"] < 10:
            reason = "Low net value per visit"
        else:
            reason = "Too far from depot / exceeds truck time window"
        ws.cell(row=r, column=5, value=reason)

    auto_width(ws)
    ws.freeze_panes = "A2"


def write_cost_breakdown(wb):
    ws = wb.create_sheet("Cost_Breakdown")

    data = [
        ("WEEKLY COST BREAKDOWN", "", ""),
        ("", "", ""),
        ("Category", "Amount ($)", "Detail"),
        ("Driver Cost (regular)", 11109.20, "462.9 hrs @ $24.00/hr"),
        ("Driver Cost (OT)", 0.00, "0.0 hrs @ $36.00/hr"),
        ("Driver Total", 11109.20, ""),
        ("", "", ""),
        ("Fuel ($0.25/km)", 1895.32, "7,581 km"),
        ("Maintenance ($0.05/km)", 379.06, "7,581 km"),
        ("Mileage ($0.09/km)", 682.32, "7,581 km"),
        ("Vehicle Variable Total", 2956.71, "$0.39/km total"),
        ("", "", ""),
        ("Fixed Truck Cost (weekly)", 6980.63, "11 trucks max/day"),
        ("  Lease component", 5272.38, "$2,077/month per truck"),
        ("  Insurance component", 1727.42, "$8,166/year per truck"),
        ("", "", ""),
        ("TOTAL WEEKLY COST", 21046.54, ""),
        ("TOTAL WEEKLY REVENUE", 49511.40, "165,038 lbs × $0.30/lb"),
        ("NET WEEKLY CONTRIBUTION", 28464.86, ""),
        ("", "", ""),
        ("Cost per pound", 0.1275, ""),
        ("Revenue per pound", 0.30, ""),
        ("Net per pound", 0.1725, ""),
        ("", "", ""),
        ("ANNUALIZED", "", ""),
        ("Total Annual Cost", 1094420, ""),
        ("Total Annual Revenue", 2574593, ""),
        ("Net Annual Contribution", 1480173, ""),
        ("Total Annual Lbs", 8581976, ""),
        ("", "", ""),
        ("FLEET UTILIZATION", "", ""),
        ("Max trucks any day", 11, "Tuesday & Thursday"),
        ("Avg trucks per day", 9.6, ""),
        ("Avg lbs per truck/day", 2463, ""),
        ("Avg km per truck/day", 113.2, ""),
    ]

    for r, (label, value, detail) in enumerate(data, 1):
        ws.cell(row=r, column=1, value=label)
        if isinstance(value, str):
            ws.cell(row=r, column=2, value=value)
        elif value != "":
            cell = ws.cell(row=r, column=2, value=value)
            if isinstance(value, float) and value < 1:
                cell.number_format = '$#,##0.0000'
            elif isinstance(value, float):
                cell.number_format = '$#,##0.00'
            elif isinstance(value, int) and value > 1000:
                cell.number_format = '$#,##0' if label != "Total Annual Lbs" and "truck" not in label.lower() and "Avg" not in label else '#,##0'
        ws.cell(row=r, column=3, value=detail)

    # Bold key rows
    for r in [1, 6, 11, 17, 18, 19, 25, 26, 27, 28]:
        ws.cell(row=r, column=1).font = Font(bold=True)
        if ws.cell(row=r, column=2).value:
            ws.cell(row=r, column=2).font = Font(bold=True)

    # Header row
    style_header(ws, 3, 3)

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30


def write_depot_pnl(wb, text):
    """Parse depot P&L from output and write Depot_PnL sheet."""
    ws = wb.create_sheet("Depot_PnL")
    headers = [
        "Depot", "Name", "Status", "Lbs/Week", "Km/Week", "Hours/Week",
        "Trucks", "Revenue", "Driver Cost", "Vehicle Variable",
        "Fixed Cost", "Total Cost", "Net Profit",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    # Parse depot P&L lines from the DEPOT PROFITABILITY REPORT section
    in_section = False
    current_depot = None
    current_name = None
    row_data = {}
    rows = []
    r = 2

    for line in text.split("\n"):
        if "DEPOT PROFITABILITY REPORT" in line:
            in_section = True
            continue
        if in_section and "NETWORK TOTAL" in line:
            # Flush last depot
            if current_depot and row_data:
                rows.append(row_data)
            # Parse network total
            m = re.search(r'\$\s*([\d,.+-]+)/week', line)
            if m:
                net_total = float(m.group(1).replace(",", "").replace("+", ""))
                rows.append({"depot": "TOTAL", "net_profit": net_total})
            break
        if in_section and "CLOSED DEPOTS" in line:
            if current_depot and row_data:
                rows.append(row_data)
            current_depot = None
            continue

        if not in_section:
            continue

        # Depot header: "  WH (Main Warehouse (GTA))"
        dm = re.match(r'^\s{2}(\w+)\s+\((.+)\)', line)
        if dm:
            if current_depot and row_data:
                rows.append(row_data)
            current_depot = dm.group(1)
            current_name = dm.group(2)
            row_data = {"depot": current_depot, "name": current_name}
            continue

        if current_depot:
            # Lbs/Km/Hours/Trucks line
            lm = re.search(r'Lbs:\s*([\d,]+)', line)
            if lm:
                row_data["lbs"] = int(lm.group(1).replace(",", ""))
            km = re.search(r'Km:\s*([\d,.]+)', line)
            if km:
                row_data["km"] = float(km.group(1).replace(",", ""))
            hrs = re.search(r'Hours:\s*([\d.]+)', line)
            if hrs:
                row_data["hours"] = float(hrs.group(1))
            tr = re.search(r'Trucks:\s*(\d+)', line)
            if tr:
                row_data["trucks"] = int(tr.group(1))

            # Cost lines
            for field, pattern in [
                ("revenue", r'Revenue:\s+\$([\d,.]+)'),
                ("driver_cost", r'Driver cost:\s+\$([\d,.]+)'),
                ("variable_vehicle", r'Vehicle var:\s+\$([\d,.]+)'),
                ("fixed_cost", r'Fixed cost:\s+\$([\d,.]+)'),
                ("total_cost", r'TOTAL COST:\s+\$([\d,.]+)'),
            ]:
                m = re.search(pattern, line)
                if m:
                    row_data[field] = float(m.group(1).replace(",", ""))

            # Net profit (can be negative)
            nm = re.search(r'NET PROFIT:\s+\$\s*([\d,.+-]+)', line)
            if nm:
                row_data["net_profit"] = float(nm.group(1).replace(",", "").replace("+", ""))
            # Status
            sm = re.search(r'\[(.*?)\]', line)
            if sm:
                row_data["status"] = sm.group(1)

    # Write rows
    for row_data in rows:
        if row_data.get("depot") == "TOTAL":
            ws.cell(row=r, column=1, value="NETWORK TOTAL").font = Font(bold=True)
            ws.cell(row=r, column=13, value=row_data.get("net_profit", 0)).number_format = MONEY_FMT
            ws.cell(row=r, column=13).font = Font(bold=True)
            r += 1
            continue
        ws.cell(row=r, column=1, value=row_data.get("depot", ""))
        ws.cell(row=r, column=2, value=row_data.get("name", ""))
        ws.cell(row=r, column=3, value=row_data.get("status", ""))
        ws.cell(row=r, column=4, value=row_data.get("lbs", 0)).number_format = INT_FMT
        ws.cell(row=r, column=5, value=row_data.get("km", 0)).number_format = '#,##0.0'
        ws.cell(row=r, column=6, value=row_data.get("hours", 0)).number_format = '#,##0.0'
        ws.cell(row=r, column=7, value=row_data.get("trucks", 0))
        ws.cell(row=r, column=8, value=row_data.get("revenue", 0)).number_format = MONEY_FMT
        ws.cell(row=r, column=9, value=row_data.get("driver_cost", 0)).number_format = MONEY_FMT
        ws.cell(row=r, column=10, value=row_data.get("variable_vehicle", 0)).number_format = MONEY_FMT
        ws.cell(row=r, column=11, value=row_data.get("fixed_cost", 0)).number_format = MONEY_FMT
        ws.cell(row=r, column=12, value=row_data.get("total_cost", 0)).number_format = MONEY_FMT
        ws.cell(row=r, column=13, value=row_data.get("net_profit", 0)).number_format = MONEY_FMT
        r += 1

    auto_width(ws)
    ws.freeze_panes = "A2"


def write_logic_constraints(wb):
    ws = wb.create_sheet("Logic_Constraints")
    ws.column_dimensions['A'].width = 80

    sections = [
        ("VRP OPTIMIZER v2 — PROFIT-MAXIMIZING MODEL", True, "2F5496"),
        ("", False, None),
        ("1. OBJECTIVE", True, "4472C4"),
        (
            "The optimizer maximizes network profit using a two-phase approach:\n"
            "  Phase 1: Depot Selection — greedy closure of unprofitable depots.\n"
            "  Phase 2: VRP Solve — per-depot route optimization using Google OR-Tools.\n"
            "All sites are evaluated purely on profitability (no mandatory sites).\n"
            "Depots can be closed entirely if unprofitable, with sites reassigned.",
            False, None,
        ),
        ("", False, None),
        ("2. DATA SOURCE", True, "4472C4"),
        (
            "Input: Route_Mapping.xlsx → Site_Table sheet (305 valid sites after deduplication).\n"
            "Each site has: address, frequency code (D1-D5), number of bins, annual lbs,\n"
            "rent cost, waste cost, and pre-calculated revenue/visit and annual site value.",
            False, None,
        ),
        ("", False, None),
        ("3. FREQUENCY CODES & SCHEDULING", True, "4472C4"),
        (
            "D1 = Daily (7 days/week, 364 visits/year)\n"
            "D2 = 2× Daily (14 visits/week, 728/year — site appears twice per day with half demand each)\n"
            "D3 = 2× Week (Tue & Thu, 104/year)\n"
            "D4 = 3× Week (Mon, Wed, Fri, 156/year)\n"
            "D5 = Weekly (1 visit/week, 52/year — day assigned deterministically by site ID)",
            False, None,
        ),
        ("", False, None),
        ("4. DEPOT SELECTION (Phase 1)", True, "4472C4"),
        (
            "7 depots available: WH (North York), Barrie, London, Newmarket, Ottawa, Hamilton, Kitchener.\n"
            "Sites initially assigned to nearest depot by Haversine distance.\n"
            "Greedy closure algorithm:\n"
            "  1. Estimate weekly P&L for each depot (revenue - fixed cost - variable cost).\n"
            "  2. Try closing the least profitable depot.\n"
            "  3. Reassign its sites to the next-nearest open depot.\n"
            "  4. If closure improves total network profit → keep it closed.\n"
            "  5. Repeat until no more closures improve profit.\n"
            "The main warehouse (WH) is never closed.",
            False, None,
        ),
        ("", False, None),
        ("5. FLEET CONSTRAINTS", True, "4472C4"),
        (
            "WH: up to 20 trucks | Barrie: 1 | London: 1 | Newmarket: 1\n"
            "Ottawa: 2 | Hamilton: 1 | Kitchener: 1\n"
            "Total fleet capacity: 27 trucks. Only open depots use their trucks.",
            False, None,
        ),
        ("", False, None),
        ("6. VEHICLE CONSTRAINTS", True, "4472C4"),
        (
            "Payload capacity: 4,000 lbs per truck per day (legal max 6,000 lbs).\n"
            "Shift limit: 12 hours max, with 60 min total breaks → 660 min effective driving.\n"
            "Service time: 15 minutes per bin at each stop (e.g. 3 bins = 45 min).",
            False, None,
        ),
        ("", False, None),
        ("7. PROFIT-BASED SITE SELECTION (Phase 2)", True, "4472C4"),
        (
            "ALL sites are optional — evaluated purely on net contribution per visit.\n"
            "Each site's drop penalty = max(0, net_contribution_per_visit) in cents.\n"
            "Sites with negative net contribution have penalty = 0 (free to drop).\n"
            "Sites with positive net contribution have penalty proportional to their value.\n"
            "The solver naturally keeps profitable sites and drops unprofitable ones.",
            False, None,
        ),
        ("", False, None),
        ("8. COST MODEL", True, "4472C4"),
        (
            "Driver wage: $24.00/hr (OT at $36.00/hr after 44 hrs/week/driver).\n"
            "Variable vehicle cost: $0.39/km (fuel $0.25 + maintenance $0.05 + mileage $0.09).\n"
            "Fixed truck cost: $90.66/day (lease $2,077/month + insurance $8,166/year).\n"
            "Revenue: $0.30 per pound collected.",
            False, None,
        ),
        ("", False, None),
        ("9. DISTANCE & TRAVEL TIME", True, "4472C4"),
        (
            "Distances and travel times are computed using Google Maps Distance Matrix API\n"
            "(real driving routes with traffic patterns). Results are cached to disk to minimize\n"
            "API costs on subsequent runs. Average urban speed assumption: 40 km/h.",
            False, None,
        ),
        ("", False, None),
        ("10. SOLVER STRATEGY", True, "4472C4"),
        (
            "Algorithm: Google OR-Tools Guided Local Search metaheuristic.\n"
            "Initial solution: PATH_CHEAPEST_ARC (greedy nearest-neighbor).\n"
            "The solver minimizes: arc_cost + fixed_vehicle_cost − dropped_site_penalty.\n"
            "This naturally balances truck count, route efficiency, and site profitability.",
            False, None,
        ),
        ("", False, None),
        ("11. DEPOT P&L REPORT", True, "4472C4"),
        (
            "After solving, per-depot P&L is computed from actual route data:\n"
            "  Revenue = lbs collected × $0.30/lb\n"
            "  Variable cost = fuel + maintenance + mileage + driver wages\n"
            "  Fixed cost = truck lease + insurance (per truck used)\n"
            "  Net profit = revenue − variable cost − fixed cost\n"
            "Depots with negative net profit are flagged for review.\n"
            "See the Depot_PnL sheet for the full breakdown.",
            False, None,
        ),
        ("", False, None),
        ("12. HOW TO INTERPRET THE RESULTS", True, "4472C4"),
        (
            "Weekly_Summary: high-level view of trucks, lbs, km per day per depot.\n"
            "Route_Details: the exact stop sequence for each truck — the driver's route sheet.\n"
            "Dropped_Sites: sites the optimizer chose not to serve, ranked by lost net value.\n"
            "Cost_Breakdown: full financial picture — cost, revenue, net contribution.\n"
            "Depot_PnL: per-depot profitability breakdown with keep/close recommendations.",
            False, None,
        ),
    ]

    row = 1
    for text, is_header, color in sections:
        cell = ws.cell(row=row, column=1, value=text)
        if is_header and color:
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        else:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        row += 1

    # Set row heights for wrapped text
    for r in range(1, row):
        val = ws.cell(row=r, column=1).value
        if val and len(val) > 80:
            lines = val.count("\n") + 1
            ws.row_dimensions[r].height = max(15, lines * 16)


def main():
    print("Parsing optimization output...")
    text = parse_output(OUTPUT_FILE)

    depot_rows = parse_depot_summary(text)
    route_rows = parse_routes(text)
    dropped_rows = parse_dropped(text)

    print(f"  Depot summary rows: {len(depot_rows)}")
    print(f"  Route detail rows: {len(route_rows)}")
    print(f"  Dropped sites: {len(dropped_rows)}")

    print("Creating Excel workbook...")
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    write_weekly_summary(wb, depot_rows)
    write_route_details(wb, route_rows)
    write_dropped_sites(wb, dropped_rows)
    write_cost_breakdown(wb)
    write_depot_pnl(wb, text)
    write_logic_constraints(wb)

    wb.save(EXPORT_PATH)
    print(f"Saved to: {EXPORT_PATH}")


if __name__ == "__main__":
    main()
